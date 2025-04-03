import script as ams
import qrcode
import openpyxl as xl

wb = xl.load_workbook('database.xlsx')
sheet = wb['Sheet1']
n = sheet.max_row


def gen_qr():
    try:
        for i in range(2, n + 1):
            regId = sheet.cell(i, 1).value
            qr_data = regId
            qr = qrcode.QRCode(version=1, box_size=10, border=1)
            qr.add_data(qr_data)
            qr.make(fit=True)
            img = qr.make_image(fill_color='#010c48', back_color='#bddbff')
            img.save(f'qrimages/{regId}.png')
        print("QR generation completed!")
    except Exception as e:
        print(str(e))


global mails_sent_number
mails_sent_number = 0


def send_mails():
    global mails_sent_number
    try:
        for i in range(2, n + 1):
            if sheet.cell(i, 2).value == "0":
                with open('conf_email.html', 'r') as file:
                    emailBody = file.read()
                # else:
                #     with open('email.html', 'r') as file:
                #         emailBody = file.read()

                service = ams.get_service()
                user_id = 'me'
                this_body = emailBody.replace('ParticipantName', sheet.cell(i, 4).value.split(' ')[0].capitalize())
                this_body = this_body.replace('imgSrc', f'cid:qrimages/{sheet.cell(i, 1).value}.png')
                msg = ams.create_message('inferencia.rkmrc@gmail.com',
                                         sheet.cell(i, 3).value,
                                         'Confirmation for Inferencia',
                                         this_body,
                                         f"qrimages/{sheet.cell(i, 1).value}.png")
                sm = ams.send_message(service, user_id, msg)
                mails_sent_number += 1
                print(f'{mails_sent_number}. Message Id: {sm["id"]} sent to {sheet.cell(i, 3).value}')

        print("All mails sent!")
    except Exception as e:
        print(str(e))


gen_qr()
